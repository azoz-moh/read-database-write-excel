import mysql.connector
import openpyxl
import pandas as pd
from datetime import datetime

"""يجلب من database ويحفظ في excel"""
Machine_time = datetime.today().strftime('%H:%M')
print(Machine_time)
# New_York_Date = datetime.now(pytz.timezone('US/Eastern')).date().strftime('%d-%m-%Y')
Machine_date = datetime.today().strftime('%d-%m-%Y')
print(Machine_date)
# هنا قم بأدخال البيانات الخاصة بقاعدة البيانات من الاسم و كلمة السر
conn = mysql.connector.connect(host='localhost', user='root', password='12345', database='gainers')
Query1 = "select Ticker, Buy, Buy_time, Shares, Sell, Sell_time from signals"
cursor = conn.cursor()
cursor.execute(Query1)
result = cursor
columns = [desc[0] for desc in result.description]
print(columns)
data = result.fetchall()
df = pd.DataFrame(list(data), columns=columns)


BuyList = []
BuyTimeList = []
SharesList = []
SellList = []
SellTimeList = []


for d in range(len(df)):
    BuyList.append(df.loc[d]['Buy'])
    BuyTimeList.append(df.loc[d]['Buy_time'])
    SharesList.append(df.loc[d]['Shares'])
    SellList.append(df.loc[d]['Sell'])
    SellTimeList.append(df.loc[d]['Sell_time'])

wb = openpyxl.load_workbook('tik2.xlsx') # اسم الملف الاكسيل
sheet1 = wb['Sheet1'] #ورقة العمل

CellIndex = []
for row in range(2, sheet1.max_row+1):
    ticker = sheet1.cell(row=row, column=1)
    Date = sheet1.cell(row=row, column=2)
    if ticker.value in df["Ticker"].values and Date.value.strftime('%d-%m-%Y')==Machine_date:
        CellIndex.append(row)


for i in range(len(CellIndex)):
    # column هو العامود الذي سيتم حفظ البيانات به
    sheet1.cell(column=8, row=CellIndex[i], value=BuyList[i])
    sheet1.cell(column=9, row=CellIndex[i], value=BuyTimeList[i])
    sheet1.cell(column=10, row=CellIndex[i], value=SharesList[i])
    sheet1.cell(column=11, row=CellIndex[i], value=(SellList[i]))
    sheet1.cell(column=12, row=CellIndex[i], value=SellTimeList[i])


wb.save('tik2.xlsx')
# Delete data from the table
conn.cursor().execute('''DELETE FROM signals
                          WHERE Checked != 0''')
conn.cursor().execute('''DELETE FROM signals
                          WHERE Checked IS NULL''')

conn.commit()
print('saved..')